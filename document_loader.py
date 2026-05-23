import csv
import json
import os
from html.parser import HTMLParser
from io import BytesIO, StringIO
from xml.etree import ElementTree as etree


TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".log",
    ".edi",
    ".x12",
    ".dat",
    ".yaml",
    ".yml",
}
STRUCTURED_EXTENSIONS = {".csv", ".json", ".html", ".htm", ".xml", ".pdf", ".docx", ".xlsx"}
SUPPORTED_EXTENSIONS = tuple(sorted(TEXT_EXTENSIONS | STRUCTURED_EXTENSIONS))
UPLOAD_TYPES = [ext.lstrip(".") for ext in SUPPORTED_EXTENSIONS]


class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    def handle_data(self, data):
        if data.strip():
            self.parts.append(data.strip())

    def text(self):
        return "\n".join(self.parts)


def is_supported_document(path_or_name):
    return os.path.splitext(path_or_name.lower())[1] in SUPPORTED_EXTENSIONS


def normalize_text(text):
    return "\n".join(line.rstrip() for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n")).strip()


def read_text_bytes(data):
    return data.decode("utf-8", errors="ignore")


def read_text_file(path):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def read_binary_file(path):
    with open(path, "rb") as f:
        return f.read()


def looks_like_text(data):
    if not data:
        return False
    sample = data[:4096]
    if sample.count(b"\x00") > 0:
        return False
    decoded = sample.decode("utf-8", errors="ignore")
    if not decoded.strip():
        return False
    printable = sum(1 for char in decoded if char.isprintable() or char.isspace())
    return printable / max(len(decoded), 1) > 0.85


def fallback_text_from_bytes(data, ext):
    if not looks_like_text(data):
        return ""
    text = read_text_bytes(data)
    return convert_text_by_extension(text, ext)


def has_expected_binary_signature(data, ext):
    if ext == ".pdf":
        return data.lstrip().startswith(b"%PDF")
    if ext in {".docx", ".xlsx"}:
        return data.startswith(b"PK")
    return True


def csv_to_text(text):
    rows = csv.reader(StringIO(text))
    return "\n".join(" | ".join(cell.strip() for cell in row) for row in rows)


def json_to_text(text):
    parsed = json.loads(text)
    return json.dumps(parsed, indent=2, ensure_ascii=False)


def html_to_text(text):
    parser = TextExtractor()
    parser.feed(text)
    return parser.text()


def xml_to_text(source):
    tree = etree.parse(source)
    root = tree.getroot()
    return etree.tostring(root, encoding="unicode")


def pdf_to_text(source):
    from pypdf import PdfReader

    reader = PdfReader(source)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def docx_to_text(source):
    from docx import Document

    doc = Document(source)
    return "\n".join(paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip())


def xlsx_to_text(source):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("Reading .xlsx files requires openpyxl. Install it with: pip install openpyxl") from e

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(value.strip() for value in values):
                rows.append(" | ".join(values))
        if rows:
            sheets.append(f"# Sheet: {sheet.title}\n" + "\n".join(rows))
    workbook.close()
    return "\n\n".join(sheets)


def yaml_to_text(text):
    try:
        import yaml
    except ImportError:
        return text

    parsed = yaml.safe_load(text)
    return json.dumps(parsed, indent=2, ensure_ascii=False)


def read_document(path):
    ext = os.path.splitext(path.lower())[1]
    if ext not in SUPPORTED_EXTENSIONS:
        return ""

    if ext in {".pdf", ".docx", ".xlsx"}:
        data = read_binary_file(path)
        if not has_expected_binary_signature(data, ext):
            return normalize_text(fallback_text_from_bytes(data, ".txt"))

        readers = {
            ".pdf": pdf_to_text,
            ".docx": docx_to_text,
            ".xlsx": xlsx_to_text,
        }
        try:
            return normalize_text(readers[ext](path))
        except Exception:
            return normalize_text(fallback_text_from_bytes(data, ".txt"))

    if ext == ".xml":
        try:
            return normalize_text(xml_to_text(path))
        except Exception:
            return normalize_text(read_text_file(path))

    text = read_text_file(path)
    return normalize_text(convert_text_by_extension(text, ext))


def read_uploaded_document(uploaded_file):
    ext = os.path.splitext(uploaded_file.name.lower())[1]
    if ext not in SUPPORTED_EXTENSIONS:
        return ""

    uploaded_file.seek(0)
    data = uploaded_file.read()

    if ext in {".pdf", ".docx", ".xlsx"}:
        if not has_expected_binary_signature(data, ext):
            return normalize_text(fallback_text_from_bytes(data, ".txt"))

        stream = BytesIO(data)
        readers = {
            ".pdf": pdf_to_text,
            ".docx": docx_to_text,
            ".xlsx": xlsx_to_text,
        }
        try:
            return normalize_text(readers[ext](stream))
        except Exception:
            return normalize_text(fallback_text_from_bytes(data, ".txt"))

    if ext == ".xml":
        stream = BytesIO(data)
        try:
            return normalize_text(xml_to_text(stream))
        except Exception:
            return normalize_text(read_text_bytes(data))

    text = read_text_bytes(data)
    return normalize_text(convert_text_by_extension(text, ext))


def convert_text_by_extension(text, ext):
    converters = {
        ".csv": csv_to_text,
        ".json": json_to_text,
        ".html": html_to_text,
        ".htm": html_to_text,
        ".yaml": yaml_to_text,
        ".yml": yaml_to_text,
    }
    return converters.get(ext, lambda value: value)(text)
